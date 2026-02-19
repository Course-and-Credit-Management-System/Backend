from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from app.core.database import get_db
from app.schemas.exam_result import ExamResultUpsertIn, ExamResultOut
from app.services.grading_service import score_to_grade
from io import BytesIO
import openpyxl


router = APIRouter(prefix="/exam-results", tags=["Admin Exam Results"])


@router.post("/import-excel")
async def import_exam_results_excel(
    file: UploadFile = File(...),
    db=Depends(get_db),
):
    if not file.filename.endswith(".xlsx"):
        return {"success": False, "message": "Only .xlsx files are supported"}

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active  # first sheet

    # Read header row
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip().lower() if cell.value else "")

    # Required columns for Enrollments collection: student_id, course_id, semesterAttend, scores
    # Optional columns: is_retake, status, grade, points, reason
    required = {"student_id", "course_id", "scores"}
    if not required.issubset(set(headers)):
        return {
            "success": False,
            "message": f"Excel must include columns: {sorted(list(required))}. Optional: semesterAttend, is_retake, status, grade, points, reason"
        }

    header_index = {h: i for i, h in enumerate(headers)}

    inserted = 0
    updated = 0
    errors = []

    for row_num in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]

        # Skip empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        try:
            student_id = str(row[header_index["student_id"]]).strip()
            course_id = str(row[header_index["course_id"]]).strip()
            scores = float(row[header_index["scores"]]) if row[header_index["scores"]] is not None else 0

            # Get optional fields
            semester_attend = None
            is_retake = False
            status = "Enrolled"
            grade = None
            points = None
            reason = None

            if "semesterattend" in header_index and row[header_index["semesterattend"]]:
                semester_attend = str(row[header_index["semesterattend"]]).strip()
            else:
                # Default to current semester if not provided
                semester_attend = "New . 1st Year . First Sem"

            if "is_retake" in header_index and row[header_index["is_retake"]] is not None:
                is_retake = bool(row[header_index["is_retake"]])

            if "status" in header_index and row[header_index["status"]]:
                status = str(row[header_index["status"]]).strip()

            # Calculate grade and points from score if not provided
            if scores is not None:
                calculated_grade, calculated_points, calculated_status = score_to_grade(scores)
                if status == "Enrolled":  # Use calculated status if default
                    status = calculated_status
                if grade is None:
                    grade = calculated_grade
                if points is None:
                    points = calculated_points

            if "grade" in header_index and row[header_index["grade"]]:
                grade = str(row[header_index["grade"]]).strip()

            if "points" in header_index and row[header_index["points"]] is not None:
                points = float(row[header_index["points"]])

            if "reason" in header_index and row[header_index["reason"]]:
                reason = str(row[header_index["reason"]]).strip()

            doc = {
                "student_id": student_id,
                "course_id": course_id,
                "semesterAttend": semester_attend,
                "scores": scores,
                "grade": grade,
                "points": points,
                "status": status,
                "is_retake": is_retake,
                "reason": reason,
            }

            # Filter for upsert - unique by student, course, semesterAttend
            filt = {
                "student_id": student_id,
                "course_id": course_id,
                "semesterAttend": semester_attend,
            }

            res = await db["Enrollments"].update_one(filt, {"$set": doc}, upsert=True)
            if res.matched_count == 0:
                inserted += 1
            else:
                updated += 1

        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})

    return {
        "success": True,
        "inserted": inserted,
        "updated": updated,
        "errors": errors,
    }
