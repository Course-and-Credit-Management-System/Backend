from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from app.core.database import get_db
from app.schemas.exam_result import ExamResultUpsertIn, ExamResultOut
from app.services.grading_service import score_to_grade
from app.services.enrollment_academic_year_service import compute_enrollment_academic_year
from io import BytesIO
import openpyxl


router = APIRouter(prefix="/exam-results", tags=["Admin Exam Results"])


def validate_year_section_major(year: int, section: str | None, major: str | None) -> tuple[str | None, str | None]:
    """
    Validate and return (section, major) based on year rules:
    - Year 1-2: section required (A/B/C), no major
    - Year 3: section required (A/B/C), major required (CS/CT)
    - Year 4-5: major required, no section
    """
    if year in [1, 2]:
        if not section or section not in ["A", "B", "C"]:
            raise ValueError(f"Section (A/B/C) is required for year {year}")
        return section, None
    elif year == 3:
        if not section or section not in ["A", "B", "C"]:
            raise ValueError("Section (A/B/C) is required for year 3")
        if not major or major not in ["CS", "CT"]:
            raise ValueError("Major (CS/CT) is required for year 3")
        return section, major
    elif year in [4, 5]:
        if not major:
            raise ValueError(f"Major is required for year {year}")
        return None, major
    else:
        raise ValueError(f"Invalid year: {year}. Must be 1-5.")


@router.post("/import-excel")
async def import_exam_results_excel(
    file: UploadFile = File(...),
    db=Depends(get_db),
):
    if not file.filename.endswith(".xlsx"):
        return {"success": False, "message": "Only .xlsx files are supported"}

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active  # first sheet

    # Read header row
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip().lower() if cell.value else "")

    # Required columns for Enrollments collection: student_id, course_id, semesterAttend, scores
    # Optional columns: is_retake, status, grade, points, reason
    required = {"student_id", "course_id", "scores"}
    if not required.issubset(set(headers)):
        return {
            "success": False,
            "message": f"Excel must include columns: {sorted(list(required))}. Optional: semesterAttend, is_retake, status, grade, points, reason"
        }

    header_index = {h: i for i, h in enumerate(headers)}

    inserted = 0
    updated = 0
    errors = []

    for row_num in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]

        # Skip empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        try:
            student_id = str(row[header_index["student_id"]]).strip()
            course_id = str(row[header_index["course_id"]]).strip()
            scores = float(row[header_index["scores"]]) if row[header_index["scores"]] is not None else 0

            # Get optional fields
            semester_attend = None
            is_retake = False
            status = "Enrolled"
            grade = None
            points = None
            reason = None

            if "semesterattend" in header_index and row[header_index["semesterattend"]]:
                semester_attend = str(row[header_index["semesterattend"]]).strip()
            else:
                # Default to current semester if not provided
                semester_attend = "New . 1st Year . First Sem"

            if "is_retake" in header_index and row[header_index["is_retake"]] is not None:
                is_retake = bool(row[header_index["is_retake"]])

            if "status" in header_index and row[header_index["status"]]:
                status = str(row[header_index["status"]]).strip()

            # Calculate grade and points from score if not provided
            if scores is not None:
                calculated_grade, calculated_points, calculated_status = score_to_grade(scores)
                # Map "Probation" status to "Failed" for Enrollments collection compatibility
                if calculated_status == "Probation":
                    calculated_status = "Failed"
                    
                if status == "Enrolled":  # Use calculated status if default
                    status = calculated_status
                if grade is None:
                    grade = calculated_grade
                if points is None:
                    points = calculated_points
                # Auto-set is_retake based on calculated status
                if calculated_status == "Failed":
                    is_retake = True

            if "grade" in header_index and row[header_index["grade"]]:
                grade = str(row[header_index["grade"]]).strip()

            if "points" in header_index and row[header_index["points"]] is not None:
                points = float(row[header_index["points"]])

            if "reason" in header_index and row[header_index["reason"]]:
                reason = str(row[header_index["reason"]]).strip()

            doc = {
                "student_id": student_id,
                "course_id": course_id,
                "semesterAttend": semester_attend,
                "academic_year": compute_enrollment_academic_year(semester_attend),
                "scores": scores,
                "grade": grade,
                "points": points,
                "status": status,
                "is_retake": is_retake,
                "reason": reason,
            }

            # Filter for upsert - unique by student, course, semesterAttend
            filt = {
                "student_id": student_id,
                "course_id": course_id,
                "semesterAttend": semester_attend,
            }

            res = await db["Enrollments"].update_one(filt, {"$set": doc}, upsert=True)
            if res.matched_count == 0:
                inserted += 1
            else:
                updated += 1

        except Exception as e:
            errors.append({"row": row_num, "error": str(e)})

    return {
        "success": True,
        "inserted": inserted,
        "updated": updated,
        "errors": errors,
    }


@router.get("")
async def list_exam_results(
    course_code: str | None = None,
    year: int | None = None,
    semester: int | None = None,
    major: str | None = None,
    section: str | None = None,
    db=Depends(get_db),
):
    """List exam results from Enrollments collection. Returns all matching records (no limit)."""
    query = {}
    
    # Map course_code to course_id for Enrollments collection
    if course_code: 
        query["course_id"] = {"$regex": course_code.strip(), "$options": "i"}
    
    # For Enrollments, we need to filter by semesterAttend which contains year/semester info
    if year is not None and year != 0:
        if semester is not None and semester != 0:
            # Create regex pattern to match semesterAttend format like "2nd Year. Second Sem"
            year_map = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th"}
            sem_map = {1: "First", 2: "Second"}
            year_str = year_map.get(year, f"{year}th")
            sem_str = sem_map.get(semester, f"{semester}")
            query["semesterAttend"] = {"$regex": f"{year_str}.*{sem_str}", "$options": "i"}
        else:
            # Filter by year only
            year_map = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th"}
            year_str = year_map.get(year, f"{year}th")
            query["semesterAttend"] = {"$regex": year_str, "$options": "i"}
    
    # Note: Enrollments collection doesn't have major/section fields, so we'll ignore these filters
    # but keep them for API compatibility

    results = await db["Enrollments"].find(query).to_list(2000)

    # Batch fetch usernames, sections, and majors to avoid N+1 query problem
    student_ids = list(set([r["student_id"] for r in results if r.get("student_id")]))
            
    users_cache = {}
    progress_cache = {}
    
    if student_ids:
        users = await db["Users"].find(
            {"user_id": {"$in": student_ids}},
            {"_id": 0, "user_id": 1, "name": 1, "student_profile.section": 1}
        ).to_list(length=None)
        
        for u in users:
            uid = u.get("user_id")
            if uid:
                sp = u.get("student_profile") or {}
                users_cache[uid] = {
                    "name": u.get("name"),
                    "section": sp.get("section")
                }
                
        # Get major from students_progress
        col_names = await db.list_collection_names()
        prog_col_name = "StudentsProgress" if "StudentsProgress" in col_names else "students_progress"
        
        progress_cursor = db[prog_col_name].find(
            {"student_id": {"$in": student_ids}},
            {"student_id": 1, "selected_major": 1}
        )
        async for pdoc in progress_cursor:
            sid = pdoc.get("student_id")
            if sid:
                progress_cache[sid] = str(pdoc.get("selected_major") or "").strip()

    out = []
    for r in results:
        r.pop("_id", None)
        uid = r.get("student_id")
        user_info = users_cache.get(uid, {})
        user_name = user_info.get("name")
        user_section = user_info.get("section")
        
        user_major = progress_cache.get(uid)
        if user_major == "" or user_major is None:
            user_major = "Unknown"
            
        # Post-filter by section and major (since Enrollments collection lacks these)
        if section and section.lower() != "all" and str(user_section).lower() != str(section).lower():
            continue
        if major and major.lower() != "all" and str(user_major).lower() != str(major).lower():
            continue
            
        # Try to infer year / semester from semesterAttend if needed for frontend map
        sa = str(r.get("semesterAttend") or "").lower()
        yr_val = 1
        sem_val = 1
        if "2nd" in sa or "second year" in sa: yr_val = 2
        elif "3rd" in sa or "third year" in sa: yr_val = 3
        elif "4th" in sa or "fourth year" in sa: yr_val = 4
        elif "5th" in sa or "fifth year" in sa: yr_val = 5
        if "second sem" in sa or "sem 2" in sa: sem_val = 2
        
        # Map Enrollments fields to ExamResults format for frontend compatibility
        mapped_result = {
            "student_id": uid,
            "student_name": user_name,
            "course_code": r["course_id"],  # Map course_id to course_code
            "year": yr_val if year is None else (year or yr_val), 
            "semester": sem_val if semester is None else (semester or sem_val),
            "section": user_section,
            "major": user_major,
            "exam_score": r.get("scores", 0),  # Map scores to exam_score
            "grade": r.get("grade", "F"),
            "grade_point": r.get("points", 0),  # Map points to grade_point
            "status": r.get("status", "Failed"),
            "is_retake": r.get("is_retake", False),  # Include the is_retake field
            "semesterAttend": r.get("semesterAttend"),  # Include for debugging
        }
        
        out.append(mapped_result)

    return out

@router.delete("")
async def delete_exam_result(
    student_id: str,
    course_code: str,
    year: int | None = None,
    semester: int | None = None,
    db=Depends(get_db),
):
    """Delete exam result from Enrollments collection. year/semester optional for malformed records."""
    enroll_col = await _get_col(db, ["Enrollments", "enrollments"])
    
    # Build filter for Enrollments collection
    filt: dict = {"student_id": student_id, "course_id": course_code}
    
    # If year and semester are provided, try multiple semesterAttend formats
    if year is not None and semester is not None:
        year_map = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th"}
        sem_map = {1: "First", 2: "Second"}
        year_str = year_map.get(year, f"{year}th")
        sem_str = sem_map.get(semester, f"{semester}")
        
        # Try different semesterAttend formats we've seen in the data
        possible_formats = [
            f"{year_str} Year. {sem_str} Sem",
            f"Old . {year_str} Year . {sem_str} Sem",
            f"New . {year_str} Year . {sem_str} Sem",
            f"{year_str} Year.{sem_str} Sem",
            f"{year_str} Year . {sem_str} Sem"
        ]
        
        for semester_attend in possible_formats:
            filt["semesterAttend"] = semester_attend
            res = await enroll_col.delete_one(filt)
            if res.deleted_count > 0:
                return {"success": True}
        
        # If none of the formats worked, try without semesterAttend
        filt_partial = {"student_id": student_id, "course_id": course_code}
        res = await enroll_col.delete_one(filt_partial)
        if res.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Exam result not found")
    else:
        # Delete without semester filter
        res = await enroll_col.delete_one(filt)
        if res.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Exam result not found")

    return {"success": True}


async def _get_col(db, names: list):
    cols = await db.list_collection_names()
    for n in names:
        if n in cols:
            return db[n]
    return db[names[0]]


@router.post("", response_model=ExamResultOut)
async def upsert_exam_result(payload: ExamResultUpsertIn, db=Depends(get_db)):
    try:
        grade, grade_point, status = score_to_grade(payload.exam_score)

        # Map "Probation" status to "Failed" for Enrollments collection compatibility
        if status == "Probation":
            status = "Failed"

        users_col = await _get_col(db, ["Users", "users"])
        enroll_col = await _get_col(db, ["Enrollments", "enrollments"])

        user_doc = await users_col.find_one({"user_id": payload.student_id})
        student_name = (user_doc or {}).get("name")

        # Create semesterAttend string from year and semester
        year_map = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th", 5: "5th"}
        sem_map = {1: "First", 2: "Second"}
        year_str = year_map.get(payload.year, f"{payload.year}th")
        sem_str = sem_map.get(payload.semester, f"{payload.semester}")
        semester_attend = f"{year_str} Year. {sem_str} Sem"

        # Map to Enrollments collection format
        doc = {
            "student_id": payload.student_id,
            "course_id": payload.course_code.upper(),  # Capitalize to maintain normalization
            "semesterAttend": semester_attend,
            "academic_year": compute_enrollment_academic_year(semester_attend),
            "status": status,
            "grade": grade,
            "points": grade_point,  # Map grade_point to points
            "scores": payload.exam_score,  # Map exam_score to scores
            "is_retake": status == "Failed",  # Set to True if student failed
        }

        # Try to find existing record first using regex for case-insensitivity AND semester match
        existing_doc = await enroll_col.find_one({
            "student_id": payload.student_id,
            "course_id": {"$regex": f"^{payload.course_code.strip()}$", "$options": "i"},
            "semesterAttend": {"$regex": f"{year_str}.*{sem_str}", "$options": "i"}
        })
        
        if existing_doc:
            # Important: Keep the original formatting to not break the frontend list regex
            doc["semesterAttend"] = existing_doc.get("semesterAttend", semester_attend)
            
            # Update existing record, preserve is_retake if already True, otherwise set based on status
            current_is_retake = existing_doc.get("is_retake", False)
            # If student failed, mark as retake (or keep existing retake status)
            new_is_retake = current_is_retake or (status == "Failed")
            
            doc["is_retake"] = new_is_retake
            res = await enroll_col.update_one(
                {"_id": existing_doc["_id"]},
                {"$set": doc}
            )
        else:
            # Create new record
            res = await enroll_col.insert_one(doc)

        # Map back to ExamResultOut format for response
        response_doc = {
            "student_id": payload.student_id,
            "student_name": student_name,
            "course_code": payload.course_code,
            "year": payload.year,
            "semester": payload.semester,
            "section": payload.section,
            "major": payload.major,
            "exam_score": payload.exam_score,
            "grade": grade,
            "grade_point": grade_point,
            "status": status,
        }

        return response_doc
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Save failed: {str(e)[:200]}")
