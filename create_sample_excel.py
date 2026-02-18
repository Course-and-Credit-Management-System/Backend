import openpyxl
from openpyxl.styles import Font, PatternFill

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Exam Results"

# Headers
headers = ["student_id", "course_id", "scores", "semesterAttend", "is_retake", "status", "grade", "points", "reason"]

# Style for headers
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Add headers with styling
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill

# Sample data
data = [
    ["TNT-1001", "CS-101", 85, "New . 1st Year . First Sem", False, "Passed", "A", 4.0, "Good performance"],
    ["TNT-1002", "CS-101", 45, "New . 1st Year . First Sem", False, "Failed", "F", 0.0, "Needs improvement"],
    ["TNT-1003", "CS-101", 72, "New . 1st Year . First Sem", True, "Passed", "B", 3.0, "Retake exam"],
    ["TNT-1004", "CS-102", 91, "New . 1st Year . First Sem", False, "Passed", "A-", 3.7, "Excellent work"],
    ["TNT-1005", "CS-102", 38, "New . 1st Year . First Sem", True, "Failed", "F", 0.0, "Second attempt"],
    ["TNT-1006", "MA-101", 88, "New . 1st Year . First Sem", False, "Passed", "B+", 3.3, "Strong performance"]
]

# Add data rows
for row_num, row_data in enumerate(data, 2):
    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=value)

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 20)
    ws.column_dimensions[column].width = adjusted_width

# Save the file
output_path = r"C:\Users\User\Desktop\Course Enroll\sample_exam_results_import.xlsx"
wb.save(output_path)

print(f"✅ Sample Excel file created: {output_path}")
print("📊 File contains 6 sample records with mixed statuses and retake flags")
