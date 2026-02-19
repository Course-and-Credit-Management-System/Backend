import asyncio
import sys
import os
import openpyxl
from io import BytesIO

# Add backend to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__))))

from app.core.config import Settings
from motor.motor_asyncio import AsyncIOMotorClient

async def test_import():
    print("Testing Excel import...")
    
    # Read the Excel file
    excel_path = r"C:\Users\User\Desktop\Course Enroll\sample_exam_results_import.xlsx"
    
    try:
        with open(excel_path, 'rb') as f:
            content = f.read()
        
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        # Read header row
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip().lower() if cell.value else "")
        
        print(f"Headers found: {headers}")
        
        # Check required columns
        required = {"student_id", "course_id", "scores"}
        missing = required - set(headers)
        if missing:
            print(f"Missing required columns: {missing}")
            return False
        else:
            print("All required columns found!")
            
        # Read first data row
        row_num = 2
        row = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"First data row: {row}")
        
        header_index = {h: i for i, h in enumerate(headers)}
        
        student_id = str(row[header_index["student_id"]]).strip()
        course_id = str(row[header_index["course_id"]]).strip()
        scores = float(row[header_index["scores"]]) if row[header_index["scores"]] is not None else 0
        
        print(f"Parsed: student_id={student_id}, course_id={course_id}, scores={scores}")
        
        return True
        
    except Exception as e:
        print(f"Error: {e}")
        return False

if __name__ == "__main__":
    success = asyncio.run(test_import())
    if success:
        print("✅ Excel file parsing works correctly")
    else:
        print("❌ Excel file parsing failed")
